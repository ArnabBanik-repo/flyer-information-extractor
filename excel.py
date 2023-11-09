import openpyxl
def saveAsExcel(filename,text_data):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in text_data:
        worksheet.append(row)
    
    excel_path = './xl_data/'+filename+'.xlsx'
    workbook.save(excel_path)
    print(filename,'saved successfully as excel.')

    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook['Sheet']

    for row in worksheet.iter_rows(values_only=True):
        for cell in row:
            print(cell, end='\t')
        print()

    workbook.close()
    return excel_path

text_data = [
    ['Name', 'Age'],
    ['Arnab', 21],
    ['Alice', 25],
    ['Bob', 35],
]


